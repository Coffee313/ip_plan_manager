from __future__ import annotations

import io

from openpyxl import Workbook, load_workbook

from app import create_app
from project_store import ProjectStore


def register(client, name: str) -> dict:
    response = client.post(
        "/api/auth/register",
        json={"login": f"user-{abs(hash(name))}"},
    )
    assert response.status_code == 200
    return response.get_json()["data"]


def user_headers(token: str) -> dict[str, str]:
    return {"X-User-Token": token}


def share_project(client, project_id: str, pin: str, owner: dict, colleague: dict) -> dict:
    invite = client.post(
        f"/api/projects/{project_id}/invite",
        headers=user_headers(owner["access_token"]),
    ).get_json()["data"]["token"]
    accepted = client.post(
        f"/api/invitations/{invite}/accept",
        json={"pin": pin},
        headers=user_headers(colleague["access_token"]),
    )
    assert accepted.status_code == 200
    return {"access_token": ""}


def workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "IP Plan"
    sheet.cell(2, 1, "10.0.0.0/16")
    sheet.cell(2, 9, "Площадка")
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_user_registration_and_restart_persistence(tmp_path):
    data_root = tmp_path / "data"
    client = create_app(ProjectStore(data_root)).test_client()

    registered = register(client, "Анна")
    token = registered["access_token"]
    assert registered["user"]["name"] == registered["user"]["login"]

    restarted = create_app(ProjectStore(data_root)).test_client()
    me = restarted.get("/api/users/me", headers=user_headers(token))
    assert me.status_code == 200
    assert me.get_json()["data"]["login"] == registered["user"]["login"]


def test_only_creator_can_delete_project(tmp_path):
    client = create_app(ProjectStore(tmp_path / "data")).test_client()
    owner = register(client, "Владелец")
    colleague = register(client, "Коллега")

    anonymous_create = client.post(
        "/api/projects", json={"name": "Без владельца", "pin": "1234"}
    )
    assert anonymous_create.status_code == 401

    created = client.post(
        "/api/projects",
        json={"name": "Проект", "pin": "1234"},
        headers=user_headers(owner["access_token"]),
    ).get_json()["data"]
    project_id = created["project"]["id"]
    owner_project_token = created["access_token"]

    owner_list = client.get(
        "/api/projects", headers=user_headers(owner["access_token"])
    ).get_json()["data"]
    colleague_list = client.get(
        "/api/projects", headers=user_headers(colleague["access_token"])
    ).get_json()["data"]
    assert owner_list[0]["can_delete"] is True
    assert colleague_list == []

    unlocked = share_project(client, project_id, "1234", owner, colleague)
    colleague_headers = {
        "X-User-Token": colleague["access_token"],
        "X-Project-Token": unlocked["access_token"],
    }
    denied = client.delete(f"/api/projects/{project_id}", headers=colleague_headers)
    assert denied.status_code == 403

    owner_headers = {
        "X-User-Token": owner["access_token"],
        "X-Project-Token": owner_project_token,
    }
    deleted = client.delete(f"/api/projects/{project_id}", headers=owner_headers)
    assert deleted.status_code == 200

    system_audit = client.get(
        "/api/system-audit",
        headers={"X-User-Token": colleague["access_token"]},
    )
    assert system_audit.status_code == 200
    event = system_audit.get_json()["data"][0]
    assert event["action"] == "project_deleted"
    assert event["project_id"] == project_id
    assert event["user_name"] == owner["user"]["login"]


def test_only_creator_can_rename_project(tmp_path):
    client = create_app(ProjectStore(tmp_path / "data")).test_client()
    owner = register(client, "Владелец")
    colleague = register(client, "Коллега")
    created = client.post(
        "/api/projects", json={"name": "Исходное имя", "pin": "1234"},
        headers=user_headers(owner["access_token"]),
    ).get_json()["data"]
    project_id = created["project"]["id"]
    unlocked = share_project(client, project_id, "1234", owner, colleague)

    denied = client.put(
        f"/api/projects/{project_id}", json={"name": "Имя коллеги"},
        headers={
            "X-User-Token": colleague["access_token"],
            "X-Project-Token": unlocked["access_token"],
        },
    )
    assert denied.status_code == 403

    allowed = client.put(
        f"/api/projects/{project_id}", json={"name": "Имя владельца"},
        headers={
            "X-User-Token": owner["access_token"],
            "X-Project-Token": created["access_token"],
        },
    )
    assert allowed.status_code == 200
    assert allowed.get_json()["data"]["name"] == "Имя владельца"


def test_only_creator_can_import_excel(tmp_path):
    client = create_app(ProjectStore(tmp_path / "data")).test_client()
    owner = register(client, "Владелец")
    colleague = register(client, "Коллега")
    created = client.post(
        "/api/projects", json={"name": "Проект", "pin": "1234"},
        headers=user_headers(owner["access_token"]),
    ).get_json()["data"]
    project_id = created["project"]["id"]
    unlocked = share_project(client, project_id, "1234", owner, colleague)

    denied = client.post(
        "/api/import",
        data={"file": (io.BytesIO(workbook_bytes()), "plan.xlsx")},
        headers={
            "X-User-Token": colleague["access_token"],
            "X-Project-ID": project_id,
            "X-Project-Token": unlocked["access_token"],
        },
    )
    assert denied.status_code == 403

    template = client.get(
        "/api/template", headers=user_headers(owner["access_token"])
    )
    workbook = load_workbook(io.BytesIO(template.data))
    sheet = workbook["IP Plan"]
    sheet.cell(2, 1, "10.0.0.0/16")
    sheet.cell(2, 9, "Площадка")
    filled_template = io.BytesIO()
    workbook.save(filled_template)
    filled_template.seek(0)

    allowed = client.post(
        "/api/import",
        data={"file": (filled_template, "plan.xlsx")},
        headers={
            "X-User-Token": owner["access_token"],
            "X-Project-ID": project_id,
            "X-Project-Token": created["access_token"],
        },
    )
    assert allowed.status_code == 200


def test_authenticated_user_can_download_empty_excel_template(tmp_path):
    client = create_app(ProjectStore(tmp_path / "data")).test_client()
    user = register(client, "Пользователь")

    response = client.get(
        "/api/template", headers=user_headers(user["access_token"])
    )

    assert response.status_code == 200
    assert response.headers["Content-Disposition"].startswith("attachment;")
    workbook = load_workbook(io.BytesIO(response.data))
    assert workbook["IP Plan"]["A1"].value == "RFC 1918"
    assert workbook["IP Plan"]["Q1"].value == "статус развертывания"