from __future__ import annotations

import io
import re
from zipfile import BadZipFile, ZipFile

import pytest
from openpyxl import Workbook

from ipplan_core import Workspace, parse_ip, parse_network
from project_store import ProjectStore


def workbook_bytes(site_cidr: str = "10.0.0.0/16") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "IP Plan"
    sheet.cell(2, 1, site_cidr)
    sheet.cell(2, 9, "Площадка")
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def normalized_xlsx_payload(data: bytes) -> dict[str, bytes]:
    with ZipFile(io.BytesIO(data)) as archive:
        payload = {name: archive.read(name) for name in archive.namelist()}
    payload["docProps/core.xml"] = re.sub(
        rb"<dcterms:modified[^>]*>.*?</dcterms:modified>",
        b"",
        payload["docProps/core.xml"],
    )
    return payload


def test_failed_import_preserves_previous_source_and_export(tmp_path):
    store = ProjectStore(tmp_path / "data")
    project, _ = store.create_project("P", "1234")
    project_id = project["id"]
    store.import_excel(project_id, workbook_bytes(), "plan.xlsx", 0)
    before, _ = store.export_excel(project_id)

    with pytest.raises(BadZipFile):
        store.import_excel(project_id, b"not an xlsx", "plan.xlsx", 1)

    after, _ = store.export_excel(project_id)
    # XLSX ZIP headers and core.xml carry the save timestamp. Compare every
    # archived payload after removing only that volatile metadata.
    assert normalized_xlsx_payload(after.getvalue()) == normalized_xlsx_payload(
        before.getvalue()
    )
    assert store.get_revision(project_id) == 1


@pytest.mark.parametrize("value", ["2001:db8::/32", "::1/128"])
def test_parse_network_rejects_ipv6(value):
    with pytest.raises(ValueError, match="IPv4"):
        parse_network(value)


@pytest.mark.parametrize("value", ["2001:db8::1", "::1"])
def test_parse_ip_rejects_ipv6(value):
    with pytest.raises(ValueError, match="IPv4"):
        parse_ip(value)


def test_corrupt_workspace_is_reported_instead_of_becoming_empty(tmp_path):
    workspace = Workspace(tmp_path)
    workspace.save()
    workspace.state_file.write_text("{broken", encoding="utf-8")

    with pytest.raises(ValueError, match="поврежден"):
        workspace.load_saved()


def test_expanding_subnet_makes_it_supernet_without_changing_neighbor(tmp_path):
    workspace = Workspace(tmp_path)
    site = workspace.create_site({"name": "Площадка", "cidr": "10.0.0.0/16"})
    container = workspace.create_subnet({"parent_id": site["id"], "cidr": "10.0.0.0/20"})
    first = workspace.create_subnet({"parent_id": container["id"], "cidr": "10.0.0.0/24"})
    second = workspace.create_subnet({"parent_id": container["id"], "cidr": "10.0.1.0/24"})

    workspace.update_subnet(first["id"], {"cidr": "10.0.0.0/23", "_changed_field": "cidr"})

    _, first_subnet = workspace.find_subnet(first["id"])
    _, second_subnet = workspace.find_subnet(second["id"])
    assert first_subnet["cidr"] == "10.0.0.0/23"
    assert second_subnet["cidr"] == "10.0.1.0/24"
    assert workspace.subnet_parent_id(workspace.sites[0], first_subnet) == container["id"]
    assert workspace.subnet_parent_id(workspace.sites[0], second_subnet) == first["id"]


def test_same_subnet_and_host_addresses_are_isolated_by_vrf(tmp_path):
    workspace = Workspace(tmp_path)
    site = workspace.create_site({"name": "Площадка", "cidr": "10.0.0.0/16"})
    blue = workspace.create_subnet({
        "parent_id": site["id"], "cidr": "10.0.1.0/24", "vrf": "BLUE",
    })
    red = workspace.create_subnet({
        "parent_id": site["id"], "cidr": "10.0.1.0/24", "vrf": "RED",
    })

    workspace.create_host(blue["id"], {"ip": "10.0.1.10", "name": "blue-srv"})
    workspace.create_host(red["id"], {"ip": "10.0.1.10", "name": "red-srv"})

    tree = workspace.state_json()["sites"][0]["tree"]
    assert [(node["cidr"], node["vrf"]) for node in tree] == [
        ("10.0.1.0/24", "BLUE"),
        ("10.0.1.0/24", "RED"),
    ]
    assert [node["hosts"][0]["name"] for node in tree] == ["blue-srv", "red-srv"]


def test_vrf_scopes_duplicate_validation_and_subnet_deletion(tmp_path):
    workspace = Workspace(tmp_path)
    site = workspace.create_site({"name": "Площадка", "cidr": "10.0.0.0/16"})
    blue = workspace.create_subnet({
        "parent_id": site["id"], "cidr": "10.0.0.0/20", "vrf": "BLUE",
    })
    workspace.create_subnet({
        "parent_id": blue["id"], "cidr": "10.0.1.0/24", "vrf": "BLUE",
    })
    red = workspace.create_subnet({
        "parent_id": site["id"], "cidr": "10.0.0.0/20", "vrf": "RED",
    })
    red_child = workspace.create_subnet({
        "parent_id": red["id"], "cidr": "10.0.1.0/24", "vrf": "RED",
    })

    with pytest.raises(ValueError, match="уже существует"):
        workspace.create_subnet({
            "parent_id": site["id"], "cidr": "10.0.0.0/20", "vrf": "RED",
        })
    with pytest.raises(ValueError, match="уже существует"):
        workspace.update_subnet(blue["id"], {
            "cidr": "10.0.0.0/20", "vrf": "RED",
        })

    workspace.delete_subnet(blue["id"])

    remaining_ids = {subnet["id"] for subnet in workspace.sites[0]["subnets"]}
    assert remaining_ids == {red["id"], red_child["id"]}


def test_new_subnet_inherits_unique_containing_vrf_and_tree_parent(tmp_path):
    workspace = Workspace(tmp_path)
    site = workspace.create_site({
        "name": "K2 Cloud", "cidr": "172.27.128.0/17",
    })
    parent = workspace.create_subnet({
        "parent_id": site["id"],
        "cidr": "172.27.128.0/21",
        "vrf": "VPC INFRA",
    })

    child = workspace.create_subnet({
        "parent_id": site["id"],
        "cidr": "172.27.134.0/24",
    })

    assert child["actual_parent_id"] == parent["id"]
    tree = workspace.state_json()["sites"][0]["tree"]
    assert [(node["cidr"], node["vrf"]) for node in tree] == [
        ("172.27.128.0/21", "VPC INFRA"),
    ]
    assert [(node["cidr"], node["vrf"]) for node in tree[0]["children"]] == [
        ("172.27.134.0/24", "VPC INFRA"),
    ]

    reloaded = Workspace(tmp_path)
    assert reloaded.load_saved() is True
    reloaded_tree = reloaded.state_json()["sites"][0]["tree"]
    assert [node["cidr"] for node in reloaded_tree[0]["children"]] == [
        "172.27.134.0/24",
    ]
