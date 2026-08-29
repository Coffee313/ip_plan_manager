from __future__ import annotations

import copy
import io
import ipaddress
import json
import uuid
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string

BASE_DIR = Path(__file__).resolve().parent
DEFAULT_DATA_DIR = BASE_DIR / "data"
SAMPLE_DIR = BASE_DIR / "sample"
DEFAULT_DATA_DIR.mkdir(exist_ok=True)

PLAN_HEADERS = [
    "RFC 1918", "Gateway", "VRF", "VLAN Number", "VLAN Name", "Comment",
    "Zone", "Site", "Subnet description", "Подсистема", "Имя сервера",
    "Серверная роль", "CPU", "RAM", "HDD/SSD", "тип оборудования",
    "статус развертывания",
]
GRAY_FILL = PatternFill(fill_type="solid", fgColor="F2F2F2")


def uid() -> str:
    return uuid.uuid4().hex


def text(value: Any) -> str:
    return "" if value is None else str(value)


def normalize_row(values: list[Any]) -> list[Any]:
    values = list(values[:17])
    if len(values) < 17:
        values += [None] * (17 - len(values))
    return values


def parse_network(value: str) -> ipaddress.IPv4Network:
    try:
        network = ipaddress.ip_network(value.strip(), strict=False)
        if not isinstance(network, ipaddress.IPv4Network):
            raise ValueError("Ожидается сеть IPv4")
        return network
    except Exception as exc:
        raise ValueError(f"Некорректный CIDR IPv4: {value}") from exc


def parse_ip(value: str) -> ipaddress.IPv4Address:
    try:
        address = ipaddress.ip_address(value.strip())
        if not isinstance(address, ipaddress.IPv4Address):
            raise ValueError("Ожидается адрес IPv4")
        return address
    except Exception as exc:
        raise ValueError(f"Некорректный IPv4-адрес: {value}") from exc


def is_cidr(value: Any) -> bool:
    if not isinstance(value, str) or "/" not in value:
        return False
    try:
        ipaddress.ip_network(value.strip(), strict=False)
        return True
    except Exception:
        return False


def is_ip(value: Any) -> bool:
    if not isinstance(value, str) or "/" in value:
        return False
    try:
        ipaddress.ip_address(value.strip())
        return True
    except Exception:
        return False


def gateway_is_usable(ip: ipaddress.IPv4Address, net: ipaddress.IPv4Network) -> bool:
    if ip not in net:
        return False
    if net.prefixlen <= 30 and ip in {net.network_address, net.broadcast_address}:
        return False
    return True


def default_gateway_for_network(net: ipaddress.IPv4Network) -> ipaddress.IPv4Address:
    if net.prefixlen == 32:
        return net.network_address
    if net.prefixlen == 31:
        return net.network_address
    return ipaddress.IPv4Address(int(net.network_address) + 1)


def gateway_for_moved_subnet(
    old_net: ipaddress.IPv4Network,
    old_gateway: ipaddress.IPv4Address | None,
    new_net: ipaddress.IPv4Network,
) -> ipaddress.IPv4Address | None:
    """Preserve gateway host offset when the subnet CIDR changes."""
    if old_gateway is None:
        return None

    offset = int(old_gateway) - int(old_net.network_address)
    candidate_num = int(new_net.network_address) + offset

    if candidate_num <= int(new_net.broadcast_address):
        candidate = ipaddress.IPv4Address(candidate_num)
        if gateway_is_usable(candidate, new_net):
            return candidate

    return default_gateway_for_network(new_net)


def subnet_for_gateway(
    prefixlen: int,
    gateway: ipaddress.IPv4Address,
) -> tuple[ipaddress.IPv4Network, ipaddress.IPv4Address]:
    """Move the subnet to the network containing the new gateway."""
    net = ipaddress.ip_network(f"{gateway}/{prefixlen}", strict=False)

    if gateway_is_usable(gateway, net):
        return net, gateway

    if prefixlen <= 30:
        if gateway == net.network_address:
            gateway = ipaddress.IPv4Address(int(net.network_address) + 1)
        elif gateway == net.broadcast_address:
            gateway = ipaddress.IPv4Address(int(net.broadcast_address) - 1)

    return net, gateway


def is_site_root(values: list[Any]) -> bool:
    # Current IP-plan convention:
    # A = site supernet, C:G empty, H empty, I = site name.
    if not is_cidr(values[0]):
        return False
    middle_empty = all(values[i] in (None, "") for i in range(2, 8))
    return middle_empty and bool(text(values[8]).strip())


def meaningful_host(values: list[Any]) -> bool:
    # Placeholder rows often contain only Site in column H.
    significant_indexes = [0, 1, 2, 3, 4, 5, 6, 8, 9, 10, 11, 12, 13, 14, 15, 16]
    return any(values[i] not in (None, "") for i in significant_indexes)


class Workspace:
    def __init__(self, data_dir: str | Path | None = None) -> None:
        self.data_dir = Path(data_dir) if data_dir is not None else DEFAULT_DATA_DIR
        self.data_dir.mkdir(parents=True, exist_ok=True)
        self.state_file = self.data_dir / "workspace.json"

        self.source_filename = ""
        self.source_path = ""
        self.source_ext = ".xlsx"
        self.sheet_name = "IP Plan"
        self.managed_start = 6
        self.style_rows = {"site": 6, "subnet": 7, "host": 8}
        self.sites: list[dict[str, Any]] = []

    # ---------- persistence ----------
    def to_dict(self) -> dict[str, Any]:
        return {
            "source_filename": self.source_filename,
            "source_path": self.source_path,
            "source_ext": self.source_ext,
            "sheet_name": self.sheet_name,
            "managed_start": self.managed_start,
            "style_rows": self.style_rows,
            "sites": self.sites,
        }

    def from_dict(self, data: dict[str, Any]) -> None:
        self.source_filename = data["source_filename"]
        self.source_path = data["source_path"]
        self.source_ext = data.get("source_ext", ".xlsx")
        self.sheet_name = data.get("sheet_name", "IP Plan")
        self.managed_start = int(data.get("managed_start", 6))
        self.style_rows = data.get("style_rows", {"site": 6, "subnet": 7, "host": 8})
        self.sites = data.get("sites", [])

    def save(self) -> None:
        self.data_dir.mkdir(parents=True, exist_ok=True)
        tmp = self.state_file.with_suffix(".json.tmp")
        tmp.write_text(
            json.dumps(self.to_dict(), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        tmp.replace(self.state_file)

    def load_saved(self) -> bool:
        if not self.state_file.exists():
            return False
        try:
            data = json.loads(self.state_file.read_text(encoding="utf-8"))
            self.from_dict(data)

            # Imported source workbooks are stored inside the project directory.
            # If an older absolute path became invalid after moving the app,
            # recover it by looking for source.xlsx/source.xlsm locally.
            if self.source_path and not Path(self.source_path).exists():
                candidate = self.data_dir / f"source{self.source_ext}"
                if candidate.exists():
                    self.source_path = str(candidate)
            return True
        except Exception as exc:
            raise ValueError(
                f"Файл состояния проекта поврежден: {self.state_file}"
            ) from exc

    # ---------- import ----------
    def import_file(self, file_bytes: bytes, filename: str) -> None:
        ext = Path(filename).suffix.lower()
        if ext not in {".xlsx", ".xlsm"}:
            raise ValueError("Поддерживаются файлы .xlsx и .xlsm")

        wb = load_workbook(
            io.BytesIO(file_bytes),
            data_only=False,
            keep_vba=(ext == ".xlsm"),
        )
        ws = wb["IP Plan"] if "IP Plan" in wb.sheetnames else wb[wb.sheetnames[0]]

        self.source_filename = filename
        self.source_ext = ext
        self.sheet_name = ws.title
        self.sites = []

        rows: list[list[Any]] = []
        for r in range(1, ws.max_row + 1):
            rows.append(normalize_row([ws.cell(r, c).value for c in range(1, 18)]))

        root_rows = [i + 1 for i, vals in enumerate(rows) if i >= 1 and is_site_root(vals)]
        if not root_rows:
            raise ValueError(
                "Не удалось определить секции площадок. "
                "Ожидается строка вида: CIDR в колонке A и название площадки в колонке I."
            )

        self.managed_start = root_rows[0]
        subnet_template = None
        host_template = None

        for pos, root_row in enumerate(root_rows):
            end_row = root_rows[pos + 1] - 1 if pos + 1 < len(root_rows) else ws.max_row
            root_values = rows[root_row - 1]
            site_name = text(root_values[8]).strip() or text(root_values[7]).strip()
            site = {
                "id": uid(),
                "cidr": str(parse_network(text(root_values[0]))),
                "name": site_name,
                "values": root_values,
                "subnets": [],
            }

            last_subnet: dict[str, Any] | None = None

            for r in range(root_row + 1, end_row + 1):
                vals = rows[r - 1]
                a = vals[0]

                if is_cidr(a):
                    if subnet_template is None:
                        subnet_template = r
                    subnet = {
                        "id": uid(),
                        "cidr": str(parse_network(text(a))),
                        "values": vals,
                        "hosts": [],
                    }
                    site["subnets"].append(subnet)
                    last_subnet = subnet
                elif meaningful_host(vals) and last_subnet is not None:
                    if host_template is None:
                        host_template = r
                    last_subnet["hosts"].append({
                        "id": uid(),
                        "values": vals,
                    })
                elif last_subnet is not None and host_template is None:
                    # First reserved detail row is a suitable style template.
                    host_template = r

            self.sites.append(site)

        self.style_rows = {
            "site": root_rows[0],
            "subnet": subnet_template or root_rows[0] + 1,
            "host": host_template or (subnet_template or root_rows[0] + 1) + 1,
        }

        # Do not replace the last known-good source until the workbook has been
        # fully parsed and validated. A malformed upload must leave export usable.
        source_path = self.data_dir / f"source{ext}"
        source_tmp = source_path.with_suffix(f"{ext}.tmp")
        source_tmp.write_bytes(file_bytes)
        source_tmp.replace(source_path)
        self.source_path = str(source_path)
        self.save()

    def load_default_if_needed(self) -> None:
        if self.load_saved():
            return
        sample = SAMPLE_DIR / "IP_Plan_Storck.xlsx"
        if sample.exists():
            self.import_file(sample.read_bytes(), sample.name)

    # ---------- lookup ----------
    def find_site(self, site_id: str) -> dict[str, Any]:
        for site in self.sites:
            if site["id"] == site_id:
                return site
        raise ValueError("Площадка не найдена")

    def find_subnet(self, subnet_id: str) -> tuple[dict[str, Any], dict[str, Any]]:
        for site in self.sites:
            for subnet in site["subnets"]:
                if subnet["id"] == subnet_id:
                    return site, subnet
        raise ValueError("Подсеть не найдена")

    def find_host(self, host_id: str) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
        for site in self.sites:
            for subnet in site["subnets"]:
                for host in subnet.get("hosts", []):
                    if host["id"] == host_id:
                        return site, subnet, host
        raise ValueError("Хост не найден")

    def find_parent_network(self, parent_id: str) -> tuple[dict[str, Any], ipaddress.IPv4Network]:
        for site in self.sites:
            if site["id"] == parent_id:
                return site, parse_network(site["cidr"])
            for subnet in site["subnets"]:
                if subnet["id"] == parent_id:
                    return site, parse_network(subnet["cidr"])
        raise ValueError("Родительская сеть не найдена")

    # ---------- hierarchy ----------
    def subnet_parent_id(self, site: dict[str, Any], subnet: dict[str, Any]) -> str:
        target = parse_network(subnet["cidr"])
        candidates: list[tuple[int, str]] = []
        for other in site["subnets"]:
            if other["id"] == subnet["id"]:
                continue
            net = parse_network(other["cidr"])
            if target.subnet_of(net) and target != net:
                candidates.append((net.prefixlen, other["id"]))
        if not candidates:
            return site["id"]
        candidates.sort(reverse=True)
        return candidates[0][1]

    def tree_for_site(self, site: dict[str, Any]) -> list[dict[str, Any]]:
        children: dict[str, list[dict[str, Any]]] = {}
        site_net = parse_network(site["cidr"])
        subnet_by_network = {
            parse_network(subnet["cidr"]): subnet for subnet in site["subnets"]
        }
        for subnet in site["subnets"]:
            target = parse_network(subnet["cidr"])
            parent_id = site["id"]
            candidate = target
            while candidate.prefixlen > site_net.prefixlen + 1:
                candidate = candidate.supernet()
                parent = subnet_by_network.get(candidate)
                if parent is not None:
                    parent_id = parent["id"]
                    break
            children.setdefault(parent_id, []).append(subnet)

        def key(node: dict[str, Any]) -> tuple[int, int]:
            n = parse_network(node["cidr"])
            return int(n.network_address), n.prefixlen

        def build(parent_id: str, depth: int = 0) -> list[dict[str, Any]]:
            nodes = []
            for subnet in sorted(children.get(parent_id, []), key=key):
                vals = normalize_row(subnet["values"])
                nodes.append({
                    "id": subnet["id"],
                    "cidr": subnet["cidr"],
                    "gateway": text(vals[1]),
                    "vrf": text(vals[2]),
                    "vlan_number": vals[3],
                    "vlan_name": text(vals[4]),
                    "comment": text(vals[5]),
                    "zone": text(vals[6]),
                    "site": text(vals[7]),
                    "description": text(vals[8]),
                    "values": vals,
                    "hosts": [self.host_json(h) for h in subnet.get("hosts", [])],
                    "children": build(subnet["id"], depth + 1),
                    "depth": depth,
                })
            return nodes

        return build(site["id"])

    def host_json(self, host: dict[str, Any]) -> dict[str, Any]:
        vals = normalize_row(host["values"])
        return {
            "id": host["id"],
            "ip": text(vals[0]),
            "comment": text(vals[5]),
            "site": text(vals[7]),
            "subsystem": text(vals[9]),
            "name": text(vals[10]),
            "role": text(vals[11]),
            "cpu": text(vals[12]),
            "ram": text(vals[13]),
            "disk": text(vals[14]),
            "type": text(vals[15]),
            "status": text(vals[16]),
            "values": vals,
        }

    def state_json(self) -> dict[str, Any]:
        return {
            "loaded": True,
            "source_filename": self.source_filename,
            "sheet_name": self.sheet_name,
            "blank_workspace": not bool(self.source_path),
            "sites": [
                {
                    "id": site["id"],
                    "cidr": site["cidr"],
                    "name": site["name"],
                    "tree": self.tree_for_site(site),
                }
                for site in self.sites
            ],
        }

    # ---------- validation ----------
    def validate_subnet_network(
        self,
        site: dict[str, Any],
        new_net: ipaddress.IPv4Network,
        exclude_id: str | None = None,
    ) -> None:
        site_net = parse_network(site["cidr"])
        if new_net == site_net or not new_net.subnet_of(site_net):
            raise ValueError(f"Подсеть должна находиться внутри {site_net}")

        for other in site["subnets"]:
            if other["id"] == exclude_id:
                continue
            old = parse_network(other["cidr"])
            if new_net == old:
                raise ValueError(f"Подсеть {new_net} уже существует")
            if new_net.overlaps(old):
                # Nested relationships are valid. Partial overlaps are not.
                if new_net.subnet_of(old) or old.subnet_of(new_net):
                    continue
                raise ValueError(f"Пересечение с существующей подсетью {old}")

    def most_specific_subnet_for_ip(
        self,
        site: dict[str, Any],
        ip: ipaddress.IPv4Address,
    ) -> dict[str, Any] | None:
        matches = []
        for subnet in site["subnets"]:
            net = parse_network(subnet["cidr"])
            if ip in net:
                matches.append((net.prefixlen, subnet))
        if not matches:
            return None
        matches.sort(key=lambda x: x[0], reverse=True)
        return matches[0][1]

    def validate_site_network(
        self,
        new_net: ipaddress.IPv4Network,
        exclude_id: str | None = None,
    ) -> None:
        for site in self.sites:
            if site["id"] == exclude_id:
                continue
            old_net = parse_network(site["cidr"])
            if new_net.overlaps(old_net):
                raise ValueError(f"Сеть площадки пересекается с {site['name']}: {old_net}")

    def create_site(self, payload: dict[str, Any]) -> dict[str, Any]:
        name = text(payload.get("name")).strip()
        if not name:
            raise ValueError("Укажите название площадки")

        net = parse_network(text(payload.get("cidr")))
        self.validate_site_network(net)

        values = [None] * 17
        values[0] = str(net)
        values[8] = name

        site = {
            "id": uid(),
            "cidr": str(net),
            "name": name,
            "values": values,
            "subnets": [],
        }
        self.sites.append(site)
        self.sites.sort(key=lambda s: (int(parse_network(s["cidr"]).network_address), parse_network(s["cidr"]).prefixlen))
        self.save()
        return {"id": site["id"]}

    def update_site(self, site_id: str, payload: dict[str, Any]) -> dict[str, Any]:
        site = self.find_site(site_id)
        old_name = site["name"]
        new_name = text(payload.get("name", old_name)).strip()
        if not new_name:
            raise ValueError("Название площадки не может быть пустым")

        new_net = parse_network(text(payload.get("cidr", site["cidr"])))
        self.validate_site_network(new_net, exclude_id=site_id)

        for subnet in site["subnets"]:
            subnet_net = parse_network(subnet["cidr"])
            if not subnet_net.subnet_of(new_net):
                raise ValueError(f"Подсеть {subnet_net} окажется вне новой сети площадки {new_net}")

        site["name"] = new_name
        site["cidr"] = str(new_net)
        values = normalize_row(site["values"])
        values[0] = str(new_net)
        values[8] = new_name
        site["values"] = values

        if new_name != old_name:
            for subnet in site["subnets"]:
                subnet_values = normalize_row(subnet["values"])
                subnet_values[7] = new_name
                subnet["values"] = subnet_values
                for host in subnet.get("hosts", []):
                    host_values = normalize_row(host["values"])
                    host_values[7] = new_name
                    host["values"] = host_values

        self.sites.sort(key=lambda s: (int(parse_network(s["cidr"]).network_address), parse_network(s["cidr"]).prefixlen))
        self.save()
        return {"cidr": str(new_net), "name": new_name}

    def delete_site(self, site_id: str) -> dict[str, int]:
        site = self.find_site(site_id)
        subnet_count = len(site["subnets"])
        host_count = sum(len(s.get("hosts", [])) for s in site["subnets"])
        self.sites = [s for s in self.sites if s["id"] != site_id]
        self.save()
        return {"subnets": subnet_count, "hosts": host_count}

    # ---------- mutations ----------
    def create_subnet(self, payload: dict[str, Any]) -> dict[str, Any]:
        parent_id = payload.get("parent_id", "")
        site, parent_net = self.find_parent_network(parent_id)
        new_net = parse_network(payload.get("cidr", ""))

        if new_net == parent_net or not new_net.subnet_of(parent_net):
            raise ValueError(f"Новая подсеть должна находиться внутри выбранной сети {parent_net}")

        self.validate_subnet_network(site, new_net)

        values = [None] * 17
        values[0] = str(new_net)
        values[1] = payload.get("gateway") or None
        values[2] = payload.get("vrf") or None
        vlan = payload.get("vlan_number")
        values[3] = int(vlan) if vlan not in (None, "") and str(vlan).strip() else None
        values[4] = payload.get("vlan_name") or None
        values[5] = payload.get("comment") or None
        values[6] = payload.get("zone") or None
        values[7] = payload.get("site") or site["name"]
        values[8] = payload.get("description") or None

        gateway = text(values[1]).strip()
        if gateway:
            gw = parse_ip(gateway)
            if gw not in new_net:
                raise ValueError("Gateway находится вне создаваемой подсети")

        subnet = {
            "id": uid(),
            "cidr": str(new_net),
            "values": values,
            "hosts": [],
        }
        site["subnets"].append(subnet)
        self.save()

        actual_parent = self.subnet_parent_id(site, subnet)
        return {"id": subnet["id"], "actual_parent_id": actual_parent}

    def update_subnet(self, subnet_id: str, payload: dict[str, Any]) -> dict[str, Any]:
        site, subnet = self.find_subnet(subnet_id)
        old_net = parse_network(subnet["cidr"])
        old_vals = normalize_row(subnet["values"])

        changed_field = text(payload.get("_changed_field")).strip()
        requested_net = parse_network(payload.get("cidr", subnet["cidr"]))
        requested_gateway_text = text(payload.get("gateway")).strip()
        old_gateway_text = text(old_vals[1]).strip()
        old_gateway = parse_ip(old_gateway_text) if old_gateway_text else None

        new_net = requested_net
        gateway: ipaddress.IPv4Address | None = (
            parse_ip(requested_gateway_text) if requested_gateway_text else None
        )
        auto_adjusted: dict[str, str] = {}

        if changed_field == "cidr":
            # CIDR is authoritative. Move Gateway by the same host offset.
            gateway = gateway_for_moved_subnet(old_net, old_gateway, new_net)
            new_gateway_text = str(gateway) if gateway is not None else ""
            if new_gateway_text != requested_gateway_text:
                auto_adjusted["gateway"] = new_gateway_text

        elif changed_field == "gateway":
            # Gateway is authoritative. Keep the existing prefix length.
            if gateway is not None:
                new_net, gateway = subnet_for_gateway(old_net.prefixlen, gateway)
                if str(new_net) != str(requested_net):
                    auto_adjusted["cidr"] = str(new_net)
                if str(gateway) != requested_gateway_text:
                    auto_adjusted["gateway"] = str(gateway)
            else:
                # Clearing Gateway leaves CIDR unchanged.
                new_net = old_net

        self.validate_subnet_network(site, new_net, exclude_id=subnet_id)

        # Hosts explicitly attached to this subnet inherit the new subnet
        # address automatically. Preserve each host's numeric offset from the
        # old network address:
        #
        # 172.17.1.0/24, host 172.17.1.10
        #        -> 172.17.2.0/24
        #        -> host 172.17.2.10
        #
        # The whole migration is validated before any host is changed.
        planned_host_ips: list[tuple[dict[str, Any], ipaddress.IPv4Address]] = []
        moved_host_ids = {h["id"] for h in subnet.get("hosts", [])}

        for host in subnet.get("hosts", []):
            ip_txt = text(host["values"][0]).strip()
            if not ip_txt:
                continue

            old_ip = parse_ip(ip_txt)
            host_offset = int(old_ip) - int(old_net.network_address)
            candidate_num = int(new_net.network_address) + host_offset

            if candidate_num > int(new_net.broadcast_address):
                raise ValueError(
                    f"Хост {old_ip} имеет смещение {host_offset} от начала сети "
                    f"{old_net}. В новой подсети {new_net} такое смещение не помещается."
                )

            new_ip = ipaddress.IPv4Address(candidate_num)

            if new_net.prefixlen <= 30 and new_ip in {
                new_net.network_address,
                new_net.broadcast_address,
            }:
                raise ValueError(
                    f"Хост {old_ip} после переноса получит недопустимый адрес "
                    f"{new_ip} (адрес сети или broadcast) в {new_net}."
                )

            # A host attached directly to the edited subnet must not land inside
            # another, more-specific subnet after the move.
            for other in site["subnets"]:
                if other["id"] == subnet_id:
                    continue
                other_net = parse_network(other["cidr"])
                if new_ip in other_net and other_net.prefixlen > new_net.prefixlen:
                    raise ValueError(
                        f"Хост {old_ip} после переноса получит {new_ip}, "
                        f"который относится к более специфичной подсети {other_net}."
                    )

            # Check for collisions with hosts outside the edited subnet.
            for other_subnet in site["subnets"]:
                for other_host in other_subnet.get("hosts", []):
                    if other_host["id"] in moved_host_ids:
                        continue
                    other_ip_txt = text(other_host["values"][0]).strip()
                    if other_ip_txt and other_ip_txt == str(new_ip):
                        raise ValueError(
                            f"Хост {old_ip} после переноса получит {new_ip}, "
                            f"но этот IP уже используется."
                        )

            planned_host_ips.append((host, new_ip))

        if gateway is not None and not gateway_is_usable(gateway, new_net):
            raise ValueError("Gateway должен быть допустимым адресом хоста внутри подсети")

        vals = normalize_row(subnet["values"])
        vals[0] = str(new_net)
        vals[1] = str(gateway) if gateway is not None else None
        vals[2] = payload.get("vrf") or None
        vlan = payload.get("vlan_number")
        vals[3] = int(vlan) if vlan not in (None, "") and str(vlan).strip() else None
        vals[4] = payload.get("vlan_name") or None
        vals[5] = payload.get("comment") or None
        vals[6] = payload.get("zone") or None
        vals[7] = payload.get("site") or site["name"]
        vals[8] = payload.get("description") or None

        subnet["cidr"] = str(new_net)
        subnet["values"] = vals

        hosts_adjusted = 0
        for host, new_ip in planned_host_ips:
            host_vals = normalize_row(host["values"])
            old_ip_txt = text(host_vals[0]).strip()
            host_vals[0] = str(new_ip)
            host["values"] = host_vals
            if old_ip_txt != str(new_ip):
                hosts_adjusted += 1

        subnet["hosts"].sort(
            key=lambda h: int(parse_ip(text(h["values"][0])))
            if text(h["values"][0]).strip() else 2**32
        )

        self.save()

        return {
            "cidr": str(new_net),
            "gateway": str(gateway) if gateway is not None else "",
            "auto_adjusted": auto_adjusted,
            "hosts_adjusted": hosts_adjusted,
        }

    def delete_subnet(self, subnet_id: str) -> dict[str, int]:
        site, subnet = self.find_subnet(subnet_id)
        target = parse_network(subnet["cidr"])

        to_delete = []
        host_count = 0
        for other in site["subnets"]:
            net = parse_network(other["cidr"])
            if net.subnet_of(target):
                to_delete.append(other["id"])
                host_count += len(other.get("hosts", []))

        site["subnets"] = [s for s in site["subnets"] if s["id"] not in set(to_delete)]
        self.save()
        return {"subnets": len(to_delete), "hosts": host_count}

    def create_host(self, subnet_id: str, payload: dict[str, Any]) -> str:
        site, subnet = self.find_subnet(subnet_id)
        net = parse_network(subnet["cidr"])
        ip = parse_ip(payload.get("ip", ""))

        if ip not in net:
            raise ValueError(f"IP должен находиться внутри {net}")
        if net.prefixlen <= 30 and ip in {net.network_address, net.broadcast_address}:
            raise ValueError("Нельзя использовать адрес сети или broadcast")

        best = self.most_specific_subnet_for_ip(site, ip)
        if best and best["id"] != subnet_id:
            raise ValueError(
                f"Адрес {ip} относится к более специфичной подсети {best['cidr']}. "
                "Добавьте хост туда."
            )

        for s in site["subnets"]:
            for h in s.get("hosts", []):
                if text(h["values"][0]).strip() == str(ip):
                    raise ValueError(f"IP {ip} уже используется")

        vals = [None] * 17
        vals[0] = str(ip)
        vals[5] = payload.get("comment") or None
        vals[7] = site["name"]
        vals[9] = payload.get("subsystem") or None
        vals[10] = payload.get("name") or None
        vals[11] = payload.get("role") or None
        vals[12] = payload.get("cpu") or None
        vals[13] = payload.get("ram") or None
        vals[14] = payload.get("disk") or None
        vals[15] = payload.get("type") or None
        vals[16] = payload.get("status") or None

        host = {"id": uid(), "values": vals}
        subnet.setdefault("hosts", []).append(host)
        subnet["hosts"].sort(
            key=lambda h: int(parse_ip(text(h["values"][0])))
            if text(h["values"][0]).strip() else 2**32
        )
        self.save()
        return host["id"]

    def update_host(self, host_id: str, payload: dict[str, Any]) -> None:
        site, subnet, host = self.find_host(host_id)
        net = parse_network(subnet["cidr"])
        ip = parse_ip(payload.get("ip", ""))

        if ip not in net:
            raise ValueError(f"IP должен находиться внутри {net}")
        if net.prefixlen <= 30 and ip in {net.network_address, net.broadcast_address}:
            raise ValueError("Нельзя использовать адрес сети или broadcast")

        best = self.most_specific_subnet_for_ip(site, ip)
        if best and best["id"] != subnet["id"]:
            raise ValueError(f"Адрес относится к более специфичной подсети {best['cidr']}")

        for s in site["subnets"]:
            for h in s.get("hosts", []):
                if h["id"] != host_id and text(h["values"][0]).strip() == str(ip):
                    raise ValueError(f"IP {ip} уже используется")

        vals = normalize_row(host["values"])
        vals[0] = str(ip)
        vals[5] = payload.get("comment") or None
        vals[7] = site["name"]
        vals[9] = payload.get("subsystem") or None
        vals[10] = payload.get("name") or None
        vals[11] = payload.get("role") or None
        vals[12] = payload.get("cpu") or None
        vals[13] = payload.get("ram") or None
        vals[14] = payload.get("disk") or None
        vals[15] = payload.get("type") or None
        vals[16] = payload.get("status") or None
        host["values"] = vals

        # Keep hosts/interfaces sorted numerically by IPv4 address after edits.
        # This mirrors create_host(), so changing e.g. .20 -> .5 immediately
        # moves the row before .10 in the web UI and later Excel export.
        subnet["hosts"].sort(
            key=lambda h: int(parse_ip(text(h["values"][0])))
            if text(h["values"][0]).strip() else 2**32
        )

        self.save()

    def delete_host(self, host_id: str) -> None:
        site, subnet, host = self.find_host(host_id)
        subnet["hosts"] = [h for h in subnet.get("hosts", []) if h["id"] != host_id]
        self.save()

    # ---------- export ----------
    def export_blank_workbook(self) -> tuple[io.BytesIO, str]:
        wb = Workbook()
        ws = wb.active
        ws.title = "IP Plan"

        thin = Side(style="thin", color="D9DEE5")
        table_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
        site_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")

        widths = {
            "A": 19, "B": 18, "C": 18, "D": 14, "E": 20, "F": 32,
            "G": 14, "H": 18, "I": 31, "J": 22, "K": 24, "L": 24,
            "M": 10, "N": 10, "O": 14, "P": 20, "Q": 23,
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        ws.sheet_view.showGridLines = False
        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.freeze_panes = "A2"

        for col, header in enumerate(PLAN_HEADERS, start=1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = copy.copy(header_fill)
            cell.border = copy.copy(table_border)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30

        current_row = 2

        # Compact summary of site supernets, matching the established template logic.
        for site in self.sites:
            ws.cell(current_row, 1, site["cidr"])
            ws.cell(current_row, 1).font = Font(bold=True)
            current_row += 1

        # One separator row between site summary and detailed sections.
        if self.sites:
            current_row += 1

        def write_row(values: list[Any], fill: PatternFill | None = None, outline_level: int = 0) -> None:
            nonlocal current_row
            vals = normalize_row(values)
            for c, value in enumerate(vals, start=1):
                cell = ws.cell(current_row, c, value)
                cell.border = copy.copy(table_border)
                if fill is not None:
                    cell.fill = copy.copy(fill)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            ws.row_dimensions[current_row].outlineLevel = min(7, outline_level)
            ws.row_dimensions[current_row].hidden = False
            current_row += 1

        def write_host_rows(node: dict[str, Any], depth: int, site_name: str) -> None:
            hosts = list(node.get("hosts", []))
            row_count = max(3, len(hosts))
            for i in range(row_count):
                vals = normalize_row(hosts[i]["values"]) if i < len(hosts) else [None] * 17
                write_row(vals, GRAY_FILL, depth)

        def write_subnet(node: dict[str, Any], depth: int, site_name: str) -> None:
            write_row(node["values"], None, depth)
            write_host_rows(node, depth + 1, site_name)
            for child in node.get("children", []):
                write_subnet(child, depth + 1, site_name)

        for site in self.sites:
            root_vals = normalize_row(site["values"])
            root_vals[0] = site["cidr"]
            root_vals[8] = site["name"]
            write_row(root_vals, site_fill, 0)

            for node in self.tree_for_site(site):
                write_subnet(node, 1, site["name"])

        last_useful_row = max(1, current_row - 1)
        ws.print_area = f"A1:Q{last_useful_row}"
        ws.column_dimensions.group("R", "XFD", outline_level=0, hidden=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output, "IP_Plan_web.xlsx"

    @classmethod
    def empty_template_bytes(cls) -> tuple[io.BytesIO, str]:
        workspace = cls.__new__(cls)
        workspace.sites = []
        return workspace.export_blank_workbook()

    def capture_row_template(self, ws, row: int) -> dict[str, Any]:
        return {
            "styles": [copy.copy(ws.cell(row, c)._style) for c in range(1, 18)],
            "height": ws.row_dimensions[row].height,
            "hidden": ws.row_dimensions[row].hidden,
            "outlineLevel": ws.row_dimensions[row].outlineLevel,
        }

    def apply_row_template(self, ws, row: int, template: dict[str, Any]) -> None:
        for c in range(1, 18):
            ws.cell(row, c)._style = copy.copy(template["styles"][c - 1])
        ws.row_dimensions[row].height = template["height"]
        ws.row_dimensions[row].hidden = False

    def export_bytes(self) -> tuple[io.BytesIO, str]:
        if not self.source_path or not Path(self.source_path).exists():
            return self.export_blank_workbook()

        keep_vba = self.source_ext == ".xlsm"
        wb = load_workbook(self.source_path, data_only=False, keep_vba=keep_vba)
        if self.sheet_name not in wb.sheetnames:
            raise ValueError(f"Лист {self.sheet_name} не найден")
        ws = wb[self.sheet_name]

        site_template = self.capture_row_template(ws, self.style_rows["site"])
        subnet_template = self.capture_row_template(ws, self.style_rows["subnet"])
        host_template = self.capture_row_template(ws, self.style_rows["host"])

        max_row = ws.max_row
        if max_row >= self.managed_start:
            ws.delete_rows(self.managed_start, max_row - self.managed_start + 1)

        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.sheet_view.showGridLines = False
        current_row = self.managed_start

        def write_values(row_num: int, values: list[Any]) -> None:
            vals = normalize_row(values)
            for col, value in enumerate(vals, start=1):
                ws.cell(row_num, col).value = value

        def write_host_rows(
            subnet: dict[str, Any],
            outline_level: int,
            site_name: str,
        ) -> None:
            nonlocal current_row
            hosts = list(subnet.get("hosts", []))
            row_count = max(3, len(hosts))
            for i in range(row_count):
                self.apply_row_template(ws, current_row, host_template)
                if i < len(hosts):
                    vals = normalize_row(hosts[i]["values"])
                else:
                    vals = [None] * 17
                write_values(current_row, vals)

                # Preserve the original template borders, including vertical
                # cell separators. Only the fill is changed to light gray.
                for c in range(1, 18):
                    cell = ws.cell(current_row, c)
                    cell.fill = copy.copy(GRAY_FILL)
                ws.row_dimensions[current_row].outlineLevel = min(7, outline_level)
                ws.row_dimensions[current_row].hidden = False
                current_row += 1

        def write_subnet(node: dict[str, Any], depth: int, site_name: str) -> None:
            nonlocal current_row
            # Subnet row.
            self.apply_row_template(ws, current_row, subnet_template)
            write_values(current_row, node["values"])
            ws.row_dimensions[current_row].outlineLevel = min(7, depth)
            ws.row_dimensions[current_row].hidden = False
            current_row += 1

            # Three host/interface slots directly under the subnet.
            write_host_rows(node, depth + 1, site_name)

            # Nested subnets are written in numeric order.
            for child in node.get("children", []):
                write_subnet(child, depth + 1, site_name)

        for site in self.sites:
            self.apply_row_template(ws, current_row, site_template)
            root_vals = normalize_row(site["values"])
            root_vals[0] = site["cidr"]
            root_vals[8] = site["name"]
            write_values(current_row, root_vals)
            ws.row_dimensions[current_row].outlineLevel = 0
            ws.row_dimensions[current_row].hidden = False
            current_row += 1

            for node in self.tree_for_site(site):
                write_subnet(node, 1, site["name"])

        # Crop the exported IP Plan to useful rows and columns only.
        #
        # The source workbook can contain formatted/colored rows and columns far
        # outside the actual IP-plan. Excel treats those style-only cells and
        # dimensions as part of the used range, which produces a large blank tail
        # after export. The web editor owns columns A:Q, so everything to the right
        # of Q and everything below the last generated row is removed.
        last_useful_row = max(1, current_row - 1)
        last_useful_col = 17  # A:Q

        # Remove merged ranges that extend outside the useful rectangle.
        for merged_range in list(ws.merged_cells.ranges):
            if merged_range.max_row > last_useful_row or merged_range.max_col > last_useful_col:
                ws.unmerge_cells(str(merged_range))

        # Delete actual trailing rows/columns.
        if ws.max_row > last_useful_row:
            ws.delete_rows(last_useful_row + 1, ws.max_row - last_useful_row)
        if ws.max_column > last_useful_col:
            ws.delete_cols(last_useful_col + 1, ws.max_column - last_useful_col)

        # Remove style-only cells that may survive row/column deletion and keep
        # Excel's used range artificially large.
        for key in list(ws._cells.keys()):
            row_idx, col_idx = key
            if row_idx > last_useful_row or col_idx > last_useful_col:
                del ws._cells[key]

        # Prune row/column dimension metadata outside the useful range.
        for row_idx in list(ws.row_dimensions.keys()):
            if isinstance(row_idx, int) and row_idx > last_useful_row:
                del ws.row_dimensions[row_idx]

        for col_key in list(ws.column_dimensions.keys()):
            try:
                col_idx = column_index_from_string(col_key)
            except Exception:
                continue
            if col_idx > last_useful_col:
                del ws.column_dimensions[col_key]

        # Limit print area as well so print/preview does not include blank space.
        ws.print_area = f"A1:Q{last_useful_row}"

        # Excel worksheets always contain columns through XFD. Hide all columns
        # after Q so the exported IP Plan visually ends at the last useful column.
        ws.column_dimensions.group("R", "XFD", outline_level=0, hidden=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        stem = Path(self.source_filename).stem or "IP_Plan"
        filename = f"{stem}_web{self.source_ext}"
        return output, filename


